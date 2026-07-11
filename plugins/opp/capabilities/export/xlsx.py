"""cap-export / xlsx — память проекта (memory/*.yaml) → xlsx-книга «Память проекта» (спека 06-f).

Read-only представление памяти (несущий принцип 2, повторяемость): одинаковая память → одинаковая
книга (детерминированный порядок листов и колонок — по контракту `schema/fact-tables.yaml`, а не
по алфавиту или порядку появления в памяти). Правки — НЕ в xlsx: через диалог с навыком → память
проекта → пересборка (`./opp export` снова). Книга не читается обратно продуктом.

Состав: лист «Титул» (название проекта + оглавление с гиперссылками на листы таблиц) и по листу
на каждую НЕПУСТУЮ таблицу памяти (имя листа = код таблицы). Колонки листа таблицы — поля контракта
в порядке YAML (`schema.model.Table.fields`); служебные ключи строки памяти вне контракта (напр.
подсписок истории правок) в колонки не попадают САМИ — просто потому, что колонки строятся из
контракта, а не из ключей строки.
"""

from __future__ import annotations
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from linter.model import load_memory
from schema.model import Schema, Table, load_schema

_MAX_COL_WIDTH = 60
_MIN_COL_WIDTH = 8

_HEADER_FILL = PatternFill("solid", fgColor="E8E8E8")
_ZEBRA_FILL = PatternFill("solid", fgColor="F6F6F6")
_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=16)
_LINK_FONT = Font(color="0563C1", underline="single")
_NOTE_FONT = Font(italic=True, size=9)
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP_TOP = Alignment(vertical="top", wrap_text=True)
_WRAP_CENTER = Alignment(vertical="center", wrap_text=True)

_PRINCIPLE_NOTE = "xlsx — представление памяти, правки — через диалог и пересборку"
_OUT_REL_PATH = ("Документы", "Память проекта.xlsx")


def _rows_by_table(workspace: Path) -> dict[str, list[dict]]:
    """Строки памяти, сгруппированные по имени таблицы (по образцу capabilities/render/engine.py)."""
    mem = load_memory(workspace)
    out: dict[str, list[dict]] = {}
    for row in mem.rows:
        out.setdefault(row.table, []).append(row.data)
    return out


def _cell_value(value):
    """Значение поля → значение ячейки: список → «; ».join, словарь → yaml-строка, иначе как есть."""
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    if isinstance(value, dict):
        return yaml.safe_dump(value, allow_unicode=True, default_flow_style=False, sort_keys=False).strip()
    return value


def _autosize(ws: Worksheet, ncols: int) -> None:
    """Грубая авто-ширина колонок по максимальной длине содержимого, с потолком."""
    for col_idx in range(1, ncols + 1):
        letter = get_column_letter(col_idx)
        max_len = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=0)
        ws.column_dimensions[letter].width = min(max(max_len + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH)


def _write_table_sheet(wb: Workbook, table: Table, rows: list[dict]) -> Worksheet:
    """Лист одной таблицы: шапка контракта (жирная, заливка, freeze_panes, автофильтр) + строки
    памяти в порядке файла, зебра чётных строк."""
    ws = wb.create_sheet(title=table.code)
    headers = [f.name for f in table.fields]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = _WRAP_CENTER

    for r_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(table.fields, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=_cell_value(row.get(field.name)))
            cell.border = _BORDER
            cell.alignment = _WRAP_TOP
            if r_idx % 2 == 0:
                cell.fill = _ZEBRA_FILL

    ws.freeze_panes = "A2"
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    _autosize(ws, len(headers))
    return ws


def _write_title_sheet(wb: Workbook, project_name: str, schema: Schema,
                        rows_by_table: dict[str, list[dict]], codes: list[str]) -> None:
    """Лист «Титул»: название проекта, оглавление (код/название/строк/гиперссылка), строка принципов."""
    ws = wb.active
    ws.title = "Титул"

    ws.cell(row=1, column=1, value=project_name).font = _TITLE_FONT
    ws.cell(row=2, column=1, value=_PRINCIPLE_NOTE).font = _NOTE_FONT

    headers = ["Код", "Название", "Строк"]
    header_row = 4
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER

    for offset, code in enumerate(codes):
        r_idx = header_row + 1 + offset
        table = schema.tables[code]
        code_cell = ws.cell(row=r_idx, column=1, value=code)
        code_cell.hyperlink = f"#'{code}'!A1"
        code_cell.font = _LINK_FONT
        ws.cell(row=r_idx, column=2, value=table.title)
        ws.cell(row=r_idx, column=3, value=len(rows_by_table.get(code) or []))
        for col_idx in range(1, 4):
            ws.cell(row=r_idx, column=col_idx).border = _BORDER

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize(ws, 3)


def export_workbook(workspace: Path) -> Path:
    """Собрать «Память проекта.xlsx» в «<workspace>/Документы/» и вернуть путь к файлу.

    Детерминированно: порядок листов и колонок — по контракту (`schema.model.load_schema`), не по
    алфавиту и не по порядку файлов памяти; лист заводится только для непустой таблицы.
    """
    workspace = Path(workspace)
    schema = load_schema()
    rows_by_table = _rows_by_table(workspace)
    # порядок контракта (dict сохраняет порядок YAML) — только непустые таблицы
    codes = [code for code in schema.tables if rows_by_table.get(code)]

    wb = Workbook()
    _write_title_sheet(wb, workspace.name or str(workspace), schema, rows_by_table, codes)
    for code in codes:
        _write_table_sheet(wb, schema.tables[code], rows_by_table[code])

    out_path = workspace.joinpath(*_OUT_REL_PATH)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path
